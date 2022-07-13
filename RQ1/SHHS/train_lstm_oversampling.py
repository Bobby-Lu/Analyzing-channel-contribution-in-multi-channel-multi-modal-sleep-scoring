import h5py
import numpy as np
from sklearn.metrics import confusion_matrix
import torch.optim as optim
from sklearn.metrics import accuracy_score
import utils
import cnn_models
import lstm_models
from torch.utils.data import Dataset
from torch.utils.data import DataLoader,SequentialSampler
import os
import torch.nn as nn
import torch
import pickle
from collections import OrderedDict
from openpyxl import Workbook
import random
import openpyxl as op
from pytorchtools import EarlyStopping

def load_cnn_model_for_lstm(model,path):
    state_dict_new_model=model.state_dict()
    checkpoint = torch.load(path,map_location=torch.device('cpu'))
    state_dict_pretrained=checkpoint['state_dict']
    state_dict_remove_module = OrderedDict()
    for k, v in state_dict_pretrained.items():
        if k!='linear.weight' and k!='linear.bias':
            #k = k[7:] # remove `module.`
            state_dict_remove_module[k] = v
    state_dict_new_model.update(state_dict_remove_module)
    model.load_state_dict(state_dict_new_model)
    return model

def loss_fn():
    criterion = nn.CrossEntropyLoss()
    return criterion

def optimizer_fn():
    optimizer = optim.Adam(lstm_model.parameters(), lr=0.0001, betas=(0.9, 0.999))
    return optimizer

def conf_mat_create(predicted,true,correct,total_30sec_epochs,conf_mat):
    total_30sec_epochs+=true.size()[0]
    correct += predicted.eq(true.view_as(predicted)).sum().item()
    conf_mat=conf_mat+confusion_matrix(true.cpu().numpy(),predicted.cpu().numpy(),labels=classes)
    return correct, total_30sec_epochs,conf_mat

def results_store_excel(correct_train,total_30sec_epochs_train,train_loss,correct_test,total_30sec_epochs_test,test_loss,epoch):
    avg_train_loss=train_loss/total_30sec_epochs_train
    avg_test_loss=test_loss/total_30sec_epochs_test
    accuracy_train=correct_train / total_30sec_epochs_train
    accuracy_test=correct_test / total_30sec_epochs_test
    lines=[epoch+1, avg_train_loss, accuracy_train, avg_test_loss, accuracy_test]
    sheet1.append(lines)

def train(cnn_model,lstm_model,data_train,data_test,batches_train,max_epochs):
    optimizer = optimizer_fn()
    early_stopping = EarlyStopping(path_to_model=path_to_lstm_model,verbose=True)
    start_epoch = 0
    lossfn = loss_fn()
    for epoch in range(start_epoch,max_epochs):
        cnn_model.eval()
        lstm_model.train()
        loss_train = 0.0
        correct_train = 0
        hidden_states_last=()
        con_mat_train = np.zeros([5,5])
        total_30sec_epochs_train=0
        batch_no = 0
        file_no=0
        count_file_len=0
        for train_idx, train_batch, train_labels in data_train:
            train_batch, train_labels=train_batch.to(device), train_labels.to(device)
            cnn_features = cnn_model(train_batch)
            decoder_input = F.one_hot(train_labels % 5, num_classes=5).type(torch.FloatTensor).to(device)
            if train_idx[0]==count_file_len:
                count_file_len+=file_length_dic_train[str(file_no)]
                file_no+=1
                print('New Patient')
                decoder_input_first = F.one_hot(torch.tensor(0) % 5, num_classes=5).type(torch.FloatTensor).to(device)
            else:
                decoder_input_first = decoder_input_last
            output_batch, decoder_input_last = lstm_model(decoder_input_first, decoder_input, cnn_features)
            train_pred = output_batch.argmax(dim=1,keepdim=True)
            train_labels_crop = train_labels.view(-1)
            loss = lossfn(output_batch,train_labels_crop)
            loss_train +=(train_labels_crop.size()[0]*loss.item())
            optimizer.zero_grad()
            loss.backward()
            optimizer.step()
            correct_train,total_30sec_epochs_train,con_mat_train=conf_mat_create(train_pred, train_labels_crop, correct_train, total_30sec_epochs_train, con_mat_train)
            print ('Epoch [{}/{}], Step [{}/{}], Loss: {:.4f}' .format(epoch+1, max_epochs, batch_no+1, batches_train, loss.item()))
            batch_no += 1
        correct_test,total_30sec_epochs_test,loss_test,con_mat_test=validation(cnn_model,lstm_model, data_test, epoch)
        results_store_excel(correct_train,total_30sec_epochs_train,loss_train,correct_test,total_30sec_epochs_test,loss_test,epoch)
        valid_loss = loss_test/total_30sec_epochs_test
        early_stopping(valid_loss,lstm_model,optimizer,epoch,con_mat_train,con_mat_test)
        if early_stopping.early_stop:
            print("Early stopping",epoch+1)
            break
    sheet2.append([0,1,2,3,4])
    for row in early_stopping.conf_mat_train_best.tolist():
        sheet2.append(row)
    sheet2.append([0,1,2,3,4])
    for row in early_stopping.conf_mat_test_best.tolist():
        sheet2.append(row)
    print('Finished Training')
        
def validation(cnn_model, lstm_model, data_test, epoch):
    cnn_model.eval()
    lstm_model.eval()
    total_30sec_epochs_test = 0
    test_loss = 0.0
    correct_test = 0
    hidden_states_last=()
    con_mat_test=np.zeros((5,5))
    lossfn1=loss_fn()
    file_no=0
    count_file_len=0
    for test_idx, test_batch, test_labels in data_test:
        test_batch, test_labels=test_batch.to(device), test_labels.to(device)
        cnn_features = cnn_model(test_batch)
        decoder_input = F.one_hot(test_labels % 5, num_classes=5).type(torch.FloatTensor).to(device)
        if test_idx[0]==count_file_len:
            count_file_len+=file_length_dic_test[str(file_no)]
            file_no+=1
            print("New Patient")
            decoder_input_first = F.one_hot(torch.tensor(0) % 5, num_classes=5).type(torch.FloatTensor).to(device)
        else:
            decoder_input_first = decoder_input_last
        output, decoder_input_last = lstm_model(decoder_input_first, decoder_input, cnn_features)
        test_labels_crop = test_labels.view(-1)
        test_pred = output.argmax(dim=1,keepdim=True)
        loss1 = lossfn1(output,test_labels_crop).item()
        test_loss += test_labels_crop.size()[0]*loss1
        correct_test,total_30sec_epochs_test,con_mat_test=conf_mat_create(test_pred,test_labels_crop,correct_test,total_30sec_epochs_test,con_mat_test)
    print("conf_mat_test:",con_mat_test)
    print("total_30sec_epochs_test:",total_30sec_epochs_test)
    print('\nTest set: Average loss: {:.4f}, Accuracy: {}/{} ({:.0f}%), Epoch:{}\n'.format(test_loss/total_30sec_epochs_test, correct_test, total_30sec_epochs_test,100. * correct_test / total_30sec_epochs_test,epoch+1))
    return correct_test,total_30sec_epochs_test,test_loss,con_mat_test


if __name__ == '__main__': 
    random.seed(30)
    batch_size = 24
    classes=[0,1,2,3,4]
    max_epochs = 200
    seq_length = 8

    path_to_cnn_model = '/RQ1/SHHS/Oversampling/models/cnn_oversampling.tar'
    path_to_lstm_model = '/RQ1/SHHS/Oversampling/models/lstm_oversampling.tar'
    
    path_to_file_length_train='/shhs/train/trainFilesNum30secEpochs_all_shhs1.pkl'
    f_file_length_train=open(path_to_file_length_train,'rb')
    file_length_dic_train=pickle.load(f_file_length_train)
    f_file_length_train.close()
    batches_train=np.sum(np.ceil(np.array(list(file_length_dic_train.values()))/(batch_size*seq_length)),dtype='int32')

    path_to_file_length_test='/shhs/val/valFilesNum30secEpochs_all_shhs1.pkl'
    f_file_length_test=open(path_to_file_length_test,'rb')
    file_length_dic_test=pickle.load(f_file_length_test)
    f_file_length_test.close()
    
    path_to_hdf5_file_train = '/shhs/train/hdf5_file_train_all_chunking_shhs1.hdf5'
    path_to_hdf5_file_test = '/shhs/val/hdf5_file_val_all_chunking_shhs1.hdf5'

    path_to_results='/RQ1/SHHS/Oversampling/results/lstm_oversampling.xlsx'
    
    if os.path.isfile(path_to_results):
        wb = op.load_workbook(path_to_results)
        sheet1 = wb['Sheet 1']
        sheet2 = wb['Sheet 2']
    else:
        wb=Workbook()
        sheet1=wb.active
        sheet1.title = "Sheet 1"
        header=['Epoch','Avg Loss Train','Accuracy Train','Avg Loss validation','Accuracy Validation']
        sheet1.append(header)
        sheet2=wb.create_sheet('Sheet 2')

    use_cuda = torch.cuda.is_available()
    device = torch.device("cuda:0" if use_cuda else "cpu")

    cnn_model = cnn_models.SleepMultiChannelNet(lstm_option=True) 
    lstm_model = lstm_models.SleepMultiChannelNet(input_size=4032,hidden_size=256,output_size=5,seq_length=seq_length,device=device)

    cnn_model = load_cnn_model_for_lstm(cnn_model,path_to_cnn_model)
    cnn_model.to(device)
    lstm_model.to(device)
    for param in cnn_model.parameters():
        param.requires_grad=False
      
    data_gen_train=utils.my_generator1(path_to_hdf5_file_train)  
    sampler_train=SequentialSampler(data_gen_train)
    batch_sampler_train=utils.CustomSequentialLSTMBatchSampler_ReturnAllChunks(sampler_train,batch_size*seq_length,file_length_dic_train,seq_length)
    data_train=DataLoader(data_gen_train,batch_size=1,num_workers=16,batch_sampler=batch_sampler_train)
    print("train dataset loaded")

    data_gen_test=utils.my_generator1(path_to_hdf5_file_test)
    sampler_test=SequentialSampler(data_gen_test)
    batch_sampler_test=utils.CustomSequentialLSTMBatchSampler_ReturnAllChunks(sampler_test,batch_size*seq_length,file_length_dic_test,seq_length)
    data_test=DataLoader(data_gen_test,batch_size=1,num_workers=16,batch_sampler=batch_sampler_test)
    print("test dataset loaded") 
     
    train(cnn_model, lstm_model, data_train, data_test, batches_train, max_epochs)
    wb.save(path_to_results)
