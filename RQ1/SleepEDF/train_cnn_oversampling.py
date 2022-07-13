import h5py    
import numpy as np  
from sklearn.metrics import confusion_matrix
import torch.optim as optim
from sklearn.metrics import accuracy_score
import utils
from cnn_models import SleepMultiChannelNet
from torch.utils.data import Dataset
from pytorchtools import EarlyStopping
from torch.utils.data import DataLoader
import os
import torch.nn as nn
import torch
import pickle
from openpyxl import Workbook
import random
import openpyxl as op

    
def loss_fn():
    criterion = nn.CrossEntropyLoss()
    return criterion
    
def optimizer_fn():
    optimizer = optim.Adam(cnn_model.parameters(), lr=0.0001, betas=(0.9, 0.999))
    return optimizer

def conf_mat_create(predicted,true,correct,total_30sec_epochs,conf_mat):
    total_30sec_epochs+=true.size()[0]
    correct += predicted.eq(true.view_as(predicted)).sum().item()
    conf_mat=conf_mat+confusion_matrix(true.cpu().numpy(),predicted.cpu().numpy(),labels=classes)
    return correct, total_30sec_epochs,conf_mat

def results_store_excel(correct_train,total_30sec_epochs_train,train_loss,correct_test,total_30sec_epochs_test,test_loss,epoch):
    avg_train_loss=train_loss/total_30sec_epochs_train
    avg_test_loss=test_loss/total_30sec_epochs_test
    accuracy_train=correct_train / total_30sec_epochs_train
    accuracy_test=correct_test / total_30sec_epochs_test
    lines=[epoch+1, avg_train_loss, accuracy_train, avg_test_loss, accuracy_test]
    sheet1.append(lines)

def train(cnn_model,data_train,data_test,batches_train,max_epochs):
    optimizer = optimizer_fn()
    early_stopping = EarlyStopping(path_to_model=path_to_cnn_model,verbose=True)
    start_epoch = 0
    lossfn = loss_fn()
    for epoch in range(start_epoch,max_epochs):
        cnn_model.train()
        loss_train = 0.0
        correct_train = 0
        con_mat_train = np.zeros([5,5])
        total_30sec_epochs_train=0
        batch_no = 0
        for train_idx, train_batch, train_labels in data_train:
            train_labels = train_labels.view(-1,1)
            train_batch, train_labels=train_batch.to(device), train_labels.to(device)
            output_batch = cnn_model(train_batch)
            train_pred = output_batch.argmax(dim=1,keepdim=True)
            train_labels_crop = train_labels.view(-1)
            loss = lossfn(output_batch,train_labels_crop)
            loss_train +=(train_labels_crop.size()[0]*loss.item())
            optimizer.zero_grad()
            loss.backward()
            optimizer.step()
            correct_train,total_30sec_epochs_train,con_mat_train=conf_mat_create(train_pred, train_labels_crop, correct_train, total_30sec_epochs_train, con_mat_train)
            print ('Epoch [{}/{}], Step [{}/{}], Loss: {:.4f}' .format(epoch+1, max_epochs, batch_no+1, batches_train, loss.item()))
            batch_no += 1
        correct_test,total_30sec_epochs_test,loss_test,con_mat_test=validation(cnn_model, data_test, epoch)
        print("total 30sec epochs in the whole training data for one epoch of training and test:",total_30sec_epochs_train,total_30sec_epochs_test)
        results_store_excel(correct_train,total_30sec_epochs_train,loss_train,correct_test,total_30sec_epochs_test,loss_test,epoch)
        valid_loss = loss_test/total_30sec_epochs_test
        early_stopping(valid_loss,cnn_model,optimizer,epoch,con_mat_train,con_mat_test)
        if early_stopping.early_stop:
            print("Early stopping",epoch+1)
            break
    sheet2.append([0,1,2,3,4])
    for row in early_stopping.conf_mat_train_best.tolist():
        sheet2.append(row)
    sheet2.append([0,1,2,3,4])
    for row in early_stopping.conf_mat_test_best.tolist():
        sheet2.append(row)
    print('Finished Training')
        

def validation(cnn_model, data_test, epoch):
    cnn_model.eval()
    total_30sec_epochs_test = 0
    test_loss = 0.0
    correct_test = 0
    con_mat_test=np.zeros((5,5))
    lossfn1=loss_fn()
    for test_idx, test_data, test_labels in data_test:
        test_data, test_labels=test_data.to(device), test_labels.to(device)
        output = cnn_model(test_data)
        test_labels_crop = test_labels.view(-1)
        test_pred = output.argmax(dim=1,keepdim=True)
        loss1 = lossfn1(output,test_labels_crop).item()
        test_loss += test_labels_crop.size()[0]*loss1
        correct_test,total_30sec_epochs_test,con_mat_test=conf_mat_create(test_pred,test_labels_crop,correct_test,total_30sec_epochs_test,con_mat_test)
    print("conf_mat_test:",con_mat_test)
    print("total_30sec_epochs_test:",total_30sec_epochs_test)
    print('\nTest set: Average loss: {:.4f}, Accuracy: {}/{} ({:.0f}%), Epoch:{}\n'.format(test_loss/total_30sec_epochs_test, correct_test, total_30sec_epochs_test,100. * correct_test / total_30sec_epochs_test,epoch+1))
    return correct_test,total_30sec_epochs_test,test_loss,con_mat_test


if __name__ == '__main__': 
    for k in range(1,21):
        for i in range(1,11):
            random.seed(30)
            batch_size = 192
            classes=[0,1,2,3,4]
            max_epochs = 200

            path_to_cnn_model = '/RQ1/SleepEDF/Oversampling/models/cnn/cnn_oversampling_outerloop'+str(k)+'_'+'innerloop'+str(i)+'.tar'
            path_to_hdf5_file_train = '/sleep-edf/20-folder-trainvalid-test/Outloop'+str(k)+'/Hdf5Files/train'+str(i-1)+'.hdf5'
            path_to_hdf5_file_test = '/sleep-edf/20-folder-trainvalid-test/Outloop'+str(k)+'/Hdf5Files/valid'+str(i-1)+'.hdf5'
            path_to_file_length_train='/sleep-edf/20-folder-trainvalid-test/Outloop'+str(k)+'/PklFiles/train'+str(i-1)+'.pkl'
            f_file_length_train=open(path_to_file_length_train,'rb')
            file_length_dic_train=pickle.load(f_file_length_train)
            batches_train=np.sum(np.ceil(np.array(list(file_length_dic_train.values()))/batch_size),dtype='int32')
            f_file_length_train.close()
            path_to_file_length_cumul='/sleep-edf/20-folder-trainvalid-test/Outloop'+str(k)+'/PklFiles/trainCumulative'+str(i-1)+'.pkl'
            f_file_length_cumul=open(path_to_file_length_cumul,'rb')
            f_file_length_dic_cumul=pickle.load(f_file_length_cumul)
            path_to_results='/RQ1/SleepEDF/Oversampling/results/cnn/cnn_oversampling_outerloop'+str(k)+'.xlsx'
            
            if os.path.isfile(path_to_results):
                wb = op.load_workbook(path_to_results)
                sheet1 = wb['Sheet 1']
                sheet2 = wb['Sheet 2']
            else:
                wb=Workbook()
                sheet1=wb.active
                sheet1.title = "Sheet 1"
                header=['Epoch','Avg Loss Train','Accuracy Train','Avg Loss validation','Accuracy Validation']
                sheet1.append(header)
                sheet2=wb.create_sheet('Sheet 2')

            use_cuda = torch.cuda.is_available()
            device = torch.device("cuda:0" if use_cuda else "cpu")

            cnn_model = SleepMultiChannelNet(lstm_option=False)
            cnn_model.to(device) 

            dataset_train=utils.my_generator1(path_to_hdf5_file_train)
            sampler=utils.CustomWeightedRandomSamplerSlicedShuffled(path_to_hdf5_file_train,f_file_length_dic_cumul)
            batch_sampler_oversampling=utils.CustomWeightedRandomBatchSamplerSlicedShuffled(sampler,batch_size,f_file_length_dic_cumul)
            batches_train=batch_sampler_oversampling.__len__()
            data_train=DataLoader(dataset_train, batch_size=1, num_workers=16, batch_sampler=batch_sampler_oversampling)

            dataset_test = utils.my_generator1(path_to_hdf5_file_test)
            data_test = DataLoader(dataset_test, batch_size=batch_size, num_workers=16)

            train(cnn_model,data_train,data_test,batches_train,max_epochs)
            wb.save(path_to_results)
        
